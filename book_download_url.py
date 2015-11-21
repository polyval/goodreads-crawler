__author__ = 'Lucien Zhou'
# -*- coding: utf-8 -*-

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import random
import time
from datetime import datetime


header = [{'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/30.0.1599.101 Safari/537.36'},
          {'User-Agent':
           'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11'},
          {'User-Agent':
           'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'},
          {'User-Agent': 'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'}]

proxies = {"http": "http://117.177.243.16:8080",
           "https": "http://117.135.250.136:82",
           }
random.seed(datetime.now())


def get_download_link(title, author):
    while 1:
        url = 'http://gen.lib.rus.ec/search.php?req='+title
        time.sleep(3)
        try:
            s = requests.session()
            response = s.post(url, headers=header[random.randint(0, 3), proxies=proxies])
            soup = BeautifulSoup(response.content)
            break
        except(requests.exceptions.URLRequired, requests.exceptions.HTTPError, requests.exceptions.ConnectionError), e:
            print e
            time.sleep(600)
            continue

    for link in soup.find_all("td"):
        if link.text in ('mobi', 'epub'):
            download_link = link.find_next("a", {"title": "Libgen"})["href"]
            return download_link
            break


wb = load_workbook(
    'C:/Users/Administrator/Desktop/genre-psychology-sociology.xlsx')
sheet1 = wb['psychology']
ws = wb.active
i = 2


while 1:
    title = sheet1['B%d' % i].value
    author = sheet1['C%d' % i].value
    if title.strip() == '' or author.strip() == '':
        break
    else:
        if get_download_link(title, author):
            ws['G%d' % i] = get_download_link(title, author)
            print 'get the ' + title+' download url'
        else:
            print 'No download url of ' + title
        i += 1
        wb.save(
            'C:/Users/Administrator/Desktop/genre-psychology-sociology.xlsx')
