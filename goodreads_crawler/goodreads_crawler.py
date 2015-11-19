__author__ = 'Lucien Zhou'
# -*- coding: utf-8 -*-

import requests
from bs4 import BeautifulSoup
import re
from openpyxl import Workbook


def book_info(genre):
    page_num = 1
    book_set = set()

    while page_num <= 100:
        url = 'https://www.goodreads.com/shelf/show/' + \
            genre+'?page='+str(page_num)

        header = {'Host': 'www.goodreads.com',
                  'Connection': 'keep-alive',
                  'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
                  'User-Agent': ' Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/36.0.1941.0 Safari/537.36',
                  'Referer': 'https://www.goodreads.com/shelf/show/sociology?page=1',
                  'Accept-Encoding': 'gzip,deflate,sdch',
                  'Accept-Language': ' zh-CN,zh;q=0.8,en;q=0.6',
                  'Cookie': 'csid=BAhJIhg0NzQtNjcxMzMzNy00NjM0NDgyBjoGRVQ%3D--658f4a016547bd627eb691ae1cb3e45e7e9bbd22; __qca=P0-672534077-1442308562125; aa_signed_outCell=in_exp; u=qTZQbGvz_AnvYgq_hGxcL0g6Z-Ivm_s7jHyD5LGDRKWmU3Z8; p=u_bRSYuFrAOJSVq8bV8RtdelpMGU4T1PNNfRWcHsq6XXHjVB; locale=en; __utmt=1; _session_id2=b68c2e32749ef8c8c05d21b3e61809de; __utma=250562704.1646777348.1442308563.1447842448.1447848289.22; __utmb=250562704.8.10.1447848289; __utmc=250562704; __utmz=250562704.1446710704.14.4.utmcsr=google|utmccn=(organic)|utmcmd=organic|utmctr=(not%20provided)'}

        try:
            s = requests.session()
            response = s.post(url, headers=header)
            soup = BeautifulSoup(response.content)
        except(requests.exceptions.URLRequired, requests.exceptions.HTTPError, requests.exceptions.ConnectionError), e:
            print e
            continue
        for book_title in soup.find_all("a", {"class": "leftAlignedImage"}):
            try:
                title = book_title["title"]
            except:
                pass
            try:
                author = book_title.find_next("span", {"itemprop": "name"})
                author_name = author.text
            except:
                pass
            try:
                rating_info = author.find_next(
                    "span", {"class": "greyText smallText"}).text
                rating = re.findall(r"rating.(.+?)\s", rating_info)[0]
                rating_number = re.findall(
                    r"(?<=\s).+?(?=ratings)", rating_info)[0]
                published_year = re.findall(
                    r"published.(.+?)\s", rating_info)[0]
            except:
                pass
            if float(rating) > 3.6 and int(filter(str.isdigit, str(rating_number))) > 20:
                book_set.add(
                    (title, author_name, rating, rating_number, published_year))
        print "Downloading Information From Page %d" % page_num
        page_num += 1
    book_list = list(book_set)
    return book_list


def mutil_genre(book_genre_lists):
    book_lists = []
    for genre in book_genre_lists:
        book_list = book_info(genre)
        book_list = sorted(book_list, key=lambda x: x[2], reverse=True)
        book_lists.append(book_list)
    return book_lists


def save_to_excel(book_lists, book_genre_lists):
    wb = Workbook(optimized_write=True)
    ws = []
    for i in range(len(book_genre_lists)):
        ws.append(wb.create_sheet(title=book_genre_lists[i]))
    for i in range(len(book_genre_lists)):
        ws[i].append(
            ['number', 'title', 'author', 'rating', 'rating_number', 'published_year'])
        count = 1
        for bl in book_lists[i]:
            ws[i].append([count, bl[0], bl[1], bl[2], bl[3], bl[4]])
            count += 1
    save_path = 'C:/Users/Administrator/Desktop/genre'
    for i in range(len(book_genre_lists)):
        save_path += ('-'+book_genre_lists[i])
    save_path += '.xlsx'
    wb.save(save_path)


book_genre_lists = ['psychology', 'sociology']
book_lists = mutil_genre(book_genre_lists)
save_to_excel(book_lists, book_genre_lists)
